import openpyxl

path = "C:\\Users\\miche\\Desktop\\.py\\ScaricaArticoli\\Order Status_2024.xlsx"

# Apri il file Excel
workbook = openpyxl.load_workbook(path)

foglio_origine = workbook.active

for riga in foglio_origine.iter_rows(min_col = 1):
    for cella in riga:
        if cella.value is not None:
            cella.value = str(cella.value).strip()

# Richiedi all'utente di inserire i numeri di ordine (separati da virgola)
input_numeri_ordine = input("Inserisci i numeri di ordine separati da virgola: ")
numeri_ordine_da_cercare = [ordine.strip() for ordine in input_numeri_ordine.split(',')]

# Crea il nuovo foglio di lavoro
foglio_destinazione = workbook.create_sheet(title='Risultato')
intestazione = [valore for valore in list(foglio_origine.iter_rows(min_row=1, max_row=1, values_only=True))[0]]
foglio_destinazione.append(intestazione)

# Cerca ciascun numero di ordine nel foglio di lavoro
for numero_ordine in numeri_ordine_da_cercare:
    trovato = False
    for riga in foglio_origine.iter_rows(min_row=2, values_only=True):
        if riga[0] == numero_ordine:
            trovato = True
            foglio_destinazione.append(riga)
            print("Operazione completata. Le righe corrispondenti sono state salvate nel nuovo foglio.")
            break

    if riga[0] != numero_ordine:
        print(f"Ordine {numero_ordine} non trovato nel foglio di lavoro.")
        
foglio_destinazione["H1"] = "Barcode"

# Salva le modifiche nel file Excel
workbook.save('tuo_file_modificato.xlsx')
